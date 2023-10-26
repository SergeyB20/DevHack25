import gdown
import openpyxl as op
from datetime import date
import datetime
import os
import pandas as pd
from datetime import datetime
import locale

with open('TEXT/manual_student.txt', 'r', encoding='utf-8') as doc:
    ManStud = doc.read()

with open('TEXT/manual_teacher.txt', 'r', encoding='utf-8') as doc:
    ManTeach = doc.read()

with open('TEXT/operating_mode.txt', 'r', encoding='utf-8') as doc:
    OperatingMode = doc.read()

with open('TEXT/bells.txt', 'r', encoding='utf-8') as doc:
    bells = doc.read()


#Нахождение четности недели
def week_check():
    global week
    now = datetime.now()
    week_number = now.weekday()

    if week_number % 7 == 0:
        week = 1
    else:
        week = 2

#Нахождение расписания для студентов на неделю из excel
def rasp():
    global n_kyrs, group
    n_kyrs = input('Введите ваш курс в виде цифры: ')
    group = input("Введите вашу группу в формате 'СПЕЦИАЛЬНОСТЬ'-'НОМЕР': ")


    global sheet_ranges
    wb = op.load_workbook(filename = 'rasp2023.xlsx')
    sheet_ranges = wb[f' {n_kyrs} курс ']

    global group_now
    i = 3
    while i < 33:
        cell_obj = sheet_ranges.cell(row = 3, column = i) 
        value = cell_obj.value
        i+=1

        if value == group:
            print('Ваша группа была найдена:', value)
            group_now = i - 1


    if week == 1:
        i = 4
        counter = 4
        while i < 88:
            if counter == 4:
                print('\n***ПОНЕДЕЛЬНИК***\n')
            if counter == 18:
                print('\n***ВТОРНИК***\n')
            if counter == 32:
                print('\n***СРЕДА***\n')
            if counter == 46:
                print('\n***ЧЕТВЕРГ***\n')
            if counter == 60:
                print('\n***ПЯТНИЦА***\n')
            if counter == 74:
                print('\n***СУББОТА***\n')
            cell_obj = sheet_ranges.cell(row = i, column = group_now) 
            lesson = cell_obj.value
            cell_obj2 = sheet_ranges.cell(row = i, column = 2) 
            time = cell_obj2.value
            if lesson is not None:
                print(time)
                print(f'{lesson}\n')
                i += 2
                counter += 2
            else:
                i += 2
                counter += 2
                 
            

    if week == 2:
        i = 5
        counter = 4
        while i < 88:
            if counter == 4:
                print('\n***ПОНЕДЕЛЬНИК***\n')
            if counter == 18:
                print('\n***ВТОРНИК***\n')
            if counter == 32:
                print('\n***СРЕДА***\n')
            if counter == 46:
                print('\n***ЧЕТВЕРГ***\n')
            if counter == 60:
                print('\n***ПЯТНИЦА***\n')
            if counter == 74:
                print('\n***СУББОТА***\n')
            cell_obj = sheet_ranges.cell(row = i, column = group_now) 
            lesson = cell_obj.value
            cell_obj2 = sheet_ranges.cell(row = (i-1), column = 2) 
            time = cell_obj2.value
            if lesson is not None:
                print(time)
                print(f'{lesson}\n')
                i += 2
                counter +=2
            else:
                i += 2
                counter +=2

#Нахождения расписания для преподавателей из excel на 2/7 дней
def prep_pars_excel():
    locale.setlocale(
        category=locale.LC_ALL,
        locale="Russian"  # чтобы дата на русском
        )

    # под каждый день недели - свой индекс, откуда начинаются пары
    days = {
        "понедельник" : 2,
        "вторник" : 9,
        "среда" : 16,
        "четверг" : 23,
        "пятница" : 30,
        "суббота" : 37
        }

        # now - текущая дата, today - день недели
    now = datetime.now()
    today = now.strftime("%A")

    #xls to xlsx ибо openpyxl с первым не работает
    try:
        os.rename("prep2023.xls", "prep2023.xlsx")
    except:
        pass

    prepod = input("ФИО препода: ")

    more = input("Вывести расписание на 2 дня или на неделю (введите 2 или 7): ")
    if not more.isdigit(): # если пользователь ввел не число - выведем ему расписание на 2 дня
        more = 2
    elif int(more) > 7: # если больше 7 - то 7
        more = 7
    elif int(more) < 1: # меньше 1 - то 2
        more = 2

    if week == 1:
        week = 0
    else: 
        week = 1

    book = op.open("prep2023.xlsx", read_only=True)
    sheet = book.worksheets[int(week)]

    for prepod_cell in range(7, 530): # диапозон строк
        cell = sheet[prepod_cell][0] # строка с преподами
        #print(cell.value)
        if cell.value == None: 
                continue
        if cell.value.replace(" ", "").lower() == prepod.replace(" ", "").lower(): # если пользователь ввел некорректно то убераем пробелы и занижаем регистр
           break # если значение ячейки - нужный нам препод, то останавливаемся

    if str(more) == '2': # выводить на 2 дня
        for i in range(days[today], days[today]+13): # от индекса 1 дня (1 пара) до конца 2 дня (7 пара) 
            group = sheet[prepod_cell+1][i].value
            time = sheet[5][i].value
            cabinet = sheet[prepod_cell+2][i].value
            day = sheet[3][i].value
            para = (i - 1) % 7 
            if para == 0:
                para = 7
            if day != None:
                print(day)
            if group == None and cabinet == None:
                print(f"{para}) -")
            else:
                print(f"{para}) Группа: {group}, аудитория: {cabinet}, время: {time}")


    elif str(more) == '7': # выводить на 7 дней
        for i in range(2, 44): # тут сразу покрываем весь диапозон дней
            group = sheet[prepod_cell+1][i].value
            time = sheet[5][i].value
            cabinet = sheet[prepod_cell+2][i].value
            day = sheet[3][i].value
            para = (i - 1) % 7
            if para == 0:
                para = 7
            if day != None:
                print(day)
            if group == None and cabinet == None:
                print(f"{para}) -")
            else:
                print(f"{para}) Группа: {group}, аудитория: {cabinet}, время: {time}")
				
#Нахождение сегоднешней даты для выгрузки файла из планшетки
def check_day():
    global Date_for_work
    Date_of_today = str(date.today())    
    Date_of_today_list = Date_of_today.split('-')
    Date_of_today_list_reversed = list(reversed(Date_of_today_list))
    Date_for_work = '.'.join(Date_of_today_list_reversed)


#Проверяет наличие планшетки
def plansрetka():
    if os.path.exists(f"Планшетка/{Date_for_work}.xlsx"): # проверяем, если планшетка уже загружена - не скачиваем
        print("Свежая планшетка уже загружена")
    else:
        url = "https://drive.google.com/drive/u/0/folders/19yyXXullGGMIT3XISiZ33wkDxHJy0zvb"
        gdown.download_folder(url, quiet=True, use_cookies=False)

#Загрузка файла планшетки для работы с ним
def load_pl():
    global book
    book = op.load_workbook(f'Планшетка/{Date_for_work}.xlsx')

#Функция поиска пар по группе в планшетке
def group_parse():
	group = input("Введите группу: ")
	group = group.lower()

	print(f"Расписание на {Date_for_work}:")
	for para in range(0, 7): # перебираем листы (пары) планшетки
		sheet = book.worksheets[para]
		for string in range(2, 39):

			if sheet[string][1].value != None: # первый ряд 
				if sheet[string][1].value.lower() == group:
					cell_group = sheet[string][1].value
					cell_cabinet = sheet[string][0].value
					cell_teacher = sheet[string][2].value
					print("Пара", para+1, cell_group)
					print(cell_cabinet, cell_teacher)

			if sheet[string][4].value != None: # второй ряд
				if sheet[string][4].value.lower() == group:
					cell_group = sheet[string][4].value
					cell_cabinet = sheet[string][3].value
					cell_teacher = sheet[string][5].value
					print("Пара", para+1, cell_group)
					print(cell_cabinet, cell_teacher)
                         
#Функция поиска пар по преподавателю в планшетке
def teacher_parse():
	teacher = input("Введите ФИО преподавателя (в формате Фамилия И.О): ")
	teacher = teacher.lower()

	print(f"Расписание на {Date_for_work}:")
	for para in range(0, 7):
		sheet = book.worksheets[para]
		for string in range(2, 39):

			if sheet[string][2].value != None:
				if sheet[string][2].value.lower() == teacher:
					cell_group = sheet[string][1].value
					cell_cabinet = sheet[string][0].value
					cell_teacher = sheet[string][2].value
					print("Пара", para+1, cell_group)
					print(cell_cabinet, cell_teacher)

			if sheet[string][5].value != None:
				if sheet[string][5].value.lower() == teacher:
					cell_group = sheet[string][4].value
					cell_cabinet = sheet[string][3].value
					cell_teacher = sheet[string][5].value
					print("Пара", para+1, cell_group)
					print(cell_cabinet, cell_teacher)